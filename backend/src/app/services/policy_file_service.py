from __future__ import annotations

from pathlib import Path
from typing import Dict, Any
from datetime import date, datetime

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[3]  # .../backend
TEMPLATE_PATH = BASE_DIR / "templates" / "policy_template.xlsx"
OUT_DIR = BASE_DIR / "generated"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def fmt_date(v: Any) -> str:
    if isinstance(v, (date, datetime)):
        return v.strftime("%d.%m.%Y")
    return str(v) if v is not None else ""


def render_policy_xlsx(values: Dict[str, Any], out_name: str) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active  # либо wb["ИмяЛиста"]

    mapping = {
        "policy_number": "B6",
        "start_date": "B9",
        "end_date": "E9",
        "insurance_days": "H9",
        "issue_date": "J9",
        "insured_full_name": "C11",
        "birthday": "H11",
        "document_number": "J11",   
        "premium_total": "H27",   
    }

    for key, cell in mapping.items():
        val = values.get(key, "")
        if key in ("start_date", "end_date", "issue_date", "birthday"):
            val = fmt_date(val)
        ws[cell].value = val

    out_path = OUT_DIR / f"{out_name}.xlsx"
    wb.save(out_path)
    return out_path
